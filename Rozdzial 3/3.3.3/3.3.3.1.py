from ad import Ad
import json
import openpyxl

ogloszenieJSON = {
            "Tytul": "",
			"Opis": "",
			"Kategoria": "",
			"Data": "",
			"Autor": "",
			"Numer_Tel": "",
			"Email": "",
			"Cena": "",
			"Wyswietlenia": ""
}


def ExcelDoJsonZapis():
    # Podaje lokalizacje pliku
    global ogloszenia
    path = "plik.xlsx"

    # Otwieram plik excela
    wb = openpyxl.load_workbook(path)

    # Zapisuje aktywny arkusz jako nowy obiekt
    sheet = wb.active

    rows = sheet.max_row
    columns = sheet.max_column

    print("Ilośc wierszy: ", rows)
    print("Ilośc kolumn: ", columns)

    for i in range(1, rows + 1):
        temp = []
        for j in range(1, columns + 1):
            cell_obj = sheet.cell(row=i, column=j)
            temp.append(cell_obj.value)

        ogloszenia.append(Ad(temp[0], temp[1], temp[2], temp[3], temp[4], temp[5], temp[6], temp[7], temp[8]))

    for ogloszenie in ogloszenia:
        with open("plik.json", "a", encoding='utf8') as file:
            ogloszenieJSON["Tytul"] = ogloszenie.title
            ogloszenieJSON["Opis"] = ogloszenie.description
            ogloszenieJSON["Kategoria"] = ogloszenie.category
            ogloszenieJSON["Data"] = ogloszenie.create_date
            ogloszenieJSON["Autor"] = ogloszenie.author
            ogloszenieJSON["Numer_Tel"] = ogloszenie.phone_number
            ogloszenieJSON["Email"] = ogloszenie.email
            ogloszenieJSON["Cena"] = ogloszenie.price
            ogloszenieJSON["Wyswietlenia"] = ogloszenie.views
            jsonObj = json.dumps(ogloszenieJSON, indent=4, default=str)
            file.write(jsonObj)