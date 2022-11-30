import openpyxl

def openExcelFile():
    #Podaje lokalizacje pliku
    global ogloszenia
    path = "plik.xlsx"

    #Otwieram plik excela
    wb = openpyxl.load_workbook(path)

    #Zapisuje aktywny arkusz jako nowy obiekt
    sheet = wb.active

    rows = sheet.max_row
    columns = sheet.max_column

    print("Ilośc wierszy: ", rows)
    print("Ilośc kolumn: ", columns)
