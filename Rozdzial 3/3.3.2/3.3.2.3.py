import openpyxl

def zapiszDoPlikuExcela():

    path = "plik.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    rows = sheet.max_row
    columns = sheet.max_column

    dane = ["Tytuł 4", "Opis 4", "Kategoria 4", "2022-09-20 00:00:00", "Autor 4", "111222333",
            "email4@email.com", 50, 300]

    for i in range(1,columns+1):
        cell = sheet.cell(row=rows+1,column=i)
        cell.value = dane[i-1]

    wb.save("plik.xlsx")
